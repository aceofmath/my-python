import openpyxl

# 새로운 엑셀 파일 생성
total_wb = openpyxl.Workbook()

# 현재 활성화된 시트 선택
total_ws = total_wb.active

# 시트 이름 변경
total_ws.title = "data"

# 헤더 추가
total_ws.append(['순번', '제품명', '가격', '수량', '합계'])

# 데이터 파일
file_list = ['11번가_1', '11번가_2', '11번가_3']

for file in file_list:
    # 데이터 불러올 엑셀 파일 열기
    wb = openpyxl.load_workbook(fr"C:\dev\workspace\py_test\resource\xlsx\{file}.xlsx", data_only=True)    
    ws = wb.active

    # 데이터 복사
    for row in ws.iter_rows(min_row = 2):
        data = []
        for cell in row:
            data.append(cell.value)
        total_ws.append(data)

# 순번 업데이트
# for row in total_ws.iter_rows(min_row=2, max_col=1):
#     for cell in row:
#         cell.value = row[0].row -1

i =0
for cell in total_ws['A']:
    if i != 0:
        cell.value = i
    i += 1

total_wb.save(r"C:\dev\workspace\py_test\resource\xlsx\11번가_통합.xlsx")