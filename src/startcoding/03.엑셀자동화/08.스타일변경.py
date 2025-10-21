import openpyxl
from openpyxl.styles import Font, Side, Border, Alignment

file_name = r"C:\dev\workspace\py_test\resource\xlsx\11번가_통합.xlsx"

# 엑셀 파일 불러오기
wb = openpyxl.load_workbook(file_name, data_only=True)

# 시트 선택
ws = wb['data']

# 형 높이 단위(범위 : 0~409, 단위 : 포인트(1포인트 = 0.035cm), 기본값 : 12.75포인트(0.42cm))
# 열 높이 단위(범위 : 0~255, 단위 : 표준 글꼴 글자 하나 크기, 기본값 : 8.43(알파벳 기준, 한 열에 8.43글자 가능))

# 열 너비 변경
ws.column_dimensions['B'].width = 30

# 행 높이 변경
ws.row_dimensions[1].height = 30

# 폰트 변경
ft = Font(size=12, color='0077FF', bold=True, italic=False, name='돋움')
ws['A1'].font = ft

# 경계선 지정
border_type1 = Side(color='000000', border_style='thin')
border_type2 = Side(color='000000', border_style='thick')

# 아래 경계선
border_bottom = Border(bottom=border_type1)
ws['A4'].border = border_bottom

# 모든 방향 경계선
border_all = Border(left=border_type2, right=border_type2, top=border_type2, bottom=border_type2)
ws['B4'].border = border_bottom

# 정렬하기
alignment_type = Alignment(horizontal='center', vertical='center', wrap_text=True)
for row in ws.iter_rows(max_row=1):
    for cell in row:
        cell.alignment = alignment_type
        cell.border = border_bottom
        cell.font = ft

# 파일 저장
wb.save(file_name)