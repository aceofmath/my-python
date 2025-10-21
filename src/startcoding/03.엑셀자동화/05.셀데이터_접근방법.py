import openpyxl

save_path= r"C:\dev\workspace\py_test\resource\xlsx\11번가.xlsx"

# 기존 엑셀 파일 불러오기
wb = openpyxl.load_workbook(save_path, data_only=True)  # data_only=True 수식이 아닌 값으로 불러오기

# data 시트 선택
ws = wb["data"]

# 01.모든 셀 데이터 가져오기
# -> 행과 열 개수를 아는 경우
# for x in range(1, 9 + 1):
#     for y in range(1, 5 + 1):
#         print(ws.cell(row=x, column=y).value, end=" ")
#     print()

# -> 행과 열 개수를 모르는 경우
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()

# 모든 행 가져오기
# for row in ws.rows:
#     print(row)

# 2번째 행부터 가져오기
# for row in ws.iter_rows(min_row=2, max_row=5):
# for row in ws.iter_rows(min_row=2, min_col=2, max_row=4, max_col=4):
for row in ws.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end=" ")
    print() 

