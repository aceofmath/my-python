import openpyxl

file_name = r"C:\dev\workspace\py_test\resource\xlsx\11번가_통합.xlsx"

# 엑셀 파일 불러오기
wb = openpyxl.load_workbook(file_name, data_only=True)

# 활성화된 시트 선택
ws = wb.active

# 제품이름 리스트
name_list = []

# 데이터 출력
for row in ws.iter_rows(min_row=2, min_col=2):
    name = row[0].value
    if name not in name_list:
        name_list.append(name)
        wb.create_sheet(name)
    data = []
    for cell in row:
        data.append(cell.value)
    wb[name].append(data)

wb.save(file_name)