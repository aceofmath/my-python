import  openpyxl  as  op  #openpyxl 모듈 import
from  openpyxl.styles  import  Alignment, PatternFill

file_name = r'C:\dev\workspace\py_test\resource\xlsx\0301_test.xlsx'

wb = op.load_workbook(file_name)

ws = wb["Sheet"]

# "C2"와 "C4"에 Text 입력  
ws["C2"].value = "Alignment test1"
ws["C4"].value = "Alignment test2"

# 셀 너비, 높이 설정하기
ws.row_dimensions[2].height = 50 #2행의 높이 50으로
ws.row_dimensions[4].height = 50 #4행의 높이 50으로
ws.column_dimensions['C'].width = 50 #C열의 너비 50으로

#Alignment test1
ws["C2"].alignment = Alignment(horizontal = 'left', vertical='center')

#Alignment test2
format1 = Alignment(horizontal = 'center', vertical='center')
ws["C4"].alignment = format1


#PatternFill test1 : green
ws["C5"].fill = PatternFill(fill_type='solid', fgColor="00FF00")

#PatternFill test2 : Black
ws["C5"].fill = PatternFill(fill_type='solid', fgColor="000000")

#Save
wb.save(file_name)
wb.close()