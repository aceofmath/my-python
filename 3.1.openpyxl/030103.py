import  openpyxl  as  op  #openpyxl 모듈 import

file_name = r'C:\dev\workspace\py_test\resource\xlsx\0301_test.xlsx'

wb = op.load_workbook(file_name)

ws = wb["Sheet5"]

#####################################################################################
# 1. Cell의 Data 읽기
#####################################################################################

#방법 1 : Sheet의 Cell 속성 사용하기
data1 = ws.cell(row=1, column=2).value

#방법 2 : 엑셀 인덱스(Range) 사용하기
data2 = ws["B1"].value

#위 결과 출력해보기
print("cell(1,2) : ", data1)
print('Range("B1"):', data2)

# 범위 이중 for문 출력
rng = ws["A1:C3"] #A1:C3 범위 저장(총 9개 Cell)

for  r_data  in  rng: #튜플의 첫번째 차원 : 색깔이 다른 영역 data
    for  c_data  in  r_data: #튜플의 두번째 차원을 위한 for문 : 1개의 Cell 요소
        print(c_data.value) #각 Cell 요소의 값 출력

#####################################################################################
# 2. Cell에 데이터 쓰기
#####################################################################################

ws = wb["Sheet"]

#"B1" Cell에 입력하기
ws.cell(row=1, column=2).value = "입력테스트1"

#"C1" Cell에 입력하기
ws["C1"].value = "입력테스트2"

# wb.save(file_name)
# wb.save(r"C:\dev\workspace\py_test\result\030103_result.xlsx") #엑셀 파일 저장(파일명 : result.xlsx)


#####################################################################################
# 3. Cell data 삭제하기
# 
# 1) cell.value 값을 공백으로 설정하기
# 2) delete_rows, delete_cols 사용하기
# 3) 시트를 통째로 삭제하고 다시 생성하기
#####################################################################################

# 1) cell.value 값을 공백으로 설정하기
ws = wb["test"]

#ws에서 데이터범위 설정
rng = ws["A1:C3"] #A1:C3 범위 저장(Cell 9개)

#튜플에 대한 for 문
for  row_data  in  rng:
    for  data  in  row_data:
        #해당 data가 2로 나눈 나머지가 0이면 공백처리
        if (data.value % 2) == 0:
            data.value = ""

#Workbook 객체 저장
# wb.save(file_name)

# 2) delete_rows, delete_cols 사용하기
#1행부터 2개행까지 행을 삭제한다.
ws.delete_rows(1,2)
#2열부터 1개열까지 열을 삭제한다.
ws.delete_cols(2,1)

#Workbook 객체 저장
# wb.save(file_name)

# 3) 시트를 통째로 삭제하고 다시 생성하기

#test 시트 삭제
wb.remove(ws)
#test 시트 재생성
wb.create_sheet('test')
wb.save(file_name)
