import  openpyxl  as  op  #openpyxl 모듈 import

file_name = r'C:\dev\workspace\py_test\resource\xlsx\0301_test.xlsx'

wb = op.load_workbook(file_name, data_only=True)

ws = wb["Sheet5"]

ws["M1"] = "합계"
ws["N1"] = "=SUM(D:D)"

wb.save(file_name)

row_max = ws.max_row  #최대 행값 저장
#for문 통해 2행~최대행까지 반복문
#range(a,b+1) : a부터 b까지 반복하는 range 구문
for  row  in  range(2, row_max+1):
    #함수 자동 작성
    ws["F"+str(row)].value = "=D"+str(row)+"*"+"E"+str(row)

wb.save(file_name)

data = [] #빈 리스트 생성

for  row  in  ws.rows:
    data.append(row[5].value)

print(data)