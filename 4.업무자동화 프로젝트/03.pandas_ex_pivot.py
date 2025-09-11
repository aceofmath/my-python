import pandas as pd
from openpyxl import load_workbook

file_name = r"C:\dev\workspace\py_test\resource\xlsx\E03EXAMPLE.xlsx"


# # 기존 시트 읽기
# df1 = pd.read_excel(file_name, sheet_name=1)
# df2 = pd.read_excel(file_name, sheet_name=2)

# # 병합
# df3 = df1.merge(df2, how="left")

# 시트 읽기 및 병합
df3 = pd.read_excel(file_name, sheet_name=1).merge(
    pd.read_excel(file_name, sheet_name=2), how="left"
)

# 기존 엑셀 불러오기
book = load_workbook(file_name)

# Sheet1
ws = book["Sheet1"] 

pdf1 = df3.pivot_table("수량", index=["업체","메뉴"], aggfunc="count")
pdf2 = df3.pivot_table("가격", index="업체", aggfunc="sum")

# with pd.ExcelWriter(file_name, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
#     pdf1.to_excel(writer, sheet_name="Sheet4", index=False)

# with pd.ExcelWriter(file_name, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
#     pdf2.to_excel(writer, sheet_name="Sheet5", index=False)

# ✅ pdf1 → Sheet1의 J14부터
start_row, start_col = 14, 10  # B2
# 헤더
for i, col_name in enumerate(pdf1.columns, start=start_col):
    ws.cell(row=start_row, column=i, value=col_name)
# 데이터
for r_idx, row in enumerate(pdf1.values, start=start_row + 1):
    for c_idx, value in enumerate(row, start=start_col):
        ws.cell(row=r_idx, column=c_idx, value=value).number_format = "#,##0"

# ✅ pdf2 → Sheet1의 I24부터
start_row, start_col = 24, 9  # C3
# 헤더
for i, col_name in enumerate(pdf2.columns, start=start_col):
    ws.cell(row=start_row, column=i, value=col_name)
# 데이터
for r_idx, row in enumerate(pdf2.values, start=start_row + 1):
    for c_idx, value in enumerate(row, start=start_col):
        ws.cell(row=r_idx, column=c_idx, value=value).number_format = "#,##0"


# pdf1.to_clipboard(index=False)
# pdf2.to_clipboard(index=False)

# 저장
book.save(file_name)

print("✅ 저장되었습니다.")